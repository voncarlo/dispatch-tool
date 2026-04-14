from __future__ import annotations

import json
import os
import re
import csv
import sqlite3
import zipfile
from datetime import datetime, timezone
from html import escape as xml_escape
from io import BytesIO
from pathlib import Path

from flask import Flask, jsonify, render_template_string, request, send_file, session
from werkzeug.security import check_password_hash, generate_password_hash


BASE_DIR = Path(__file__).resolve().parent
HTML_FILE = BASE_DIR / "dispatch_tool.html"
DB_FILE = BASE_DIR / "activity_logs.db"
OWNER_PASSWORD = os.environ.get("DISPATCH_OWNER_PASSWORD", "Torero@2026")
PHONE_LIST_DSP_KEYS = ("armm", "tlc", "portkey", "mstar")
DSP_NAMES = {
    "armm": "ARMM Logistics",
    "tlc": "Torero Logistics Corp",
    "portkey": "PortKey Delivery",
    "mstar": "MSTAR Shipping",
    "hrt": "HR Transportation",
}
ASSOCIATE_DATA_FILES = {
    "portkey": (
        BASE_DIR / "AssociateData.csv",
        Path.home() / "Downloads" / "AssociateData.csv",
    )
}
VEHICLE_DATA_FILES = {
    "portkey": (
        BASE_DIR / "VehiclesData.xlsx",
        Path.home() / "Downloads" / "VehiclesData (4).xlsx",
    )
}
DVIC_PAPER_TEMPLATE_FILES = (
    BASE_DIR / "DVIC Paper Inspection.docx",
    Path(r"z:\08 - Clients\Clients - LMS - PRKL\01 - Dispatch\PRKL - DVIC\DVIC Paper Inspection.docx"),
)

app = Flask(__name__)
app.secret_key = os.environ.get("DISPATCH_SECRET_KEY", f"{OWNER_PASSWORD}-dispatch-session")


def get_db_connection() -> sqlite3.Connection:
    conn = sqlite3.connect(DB_FILE)
    conn.row_factory = sqlite3.Row
    return conn


def init_db() -> None:
    with get_db_connection() as conn:
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS phone_list_entries (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                dsp_key TEXT NOT NULL,
                label TEXT NOT NULL,
                last_name TEXT,
                work_phone TEXT,
                home_phone TEXT,
                mobile_phone TEXT,
                updated_at TEXT NOT NULL
            )
            """
        )
        conn.execute(
            "CREATE INDEX IF NOT EXISTS idx_phone_list_entries_dsp_key ON phone_list_entries (dsp_key)"
        )
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS transporter_id_entries (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                dsp_key TEXT NOT NULL,
                da_key TEXT NOT NULL,
                da_name TEXT NOT NULL,
                transporter_id TEXT NOT NULL,
                notes TEXT,
                updated_at TEXT NOT NULL,
                UNIQUE (dsp_key, da_key)
            )
            """
        )
        conn.execute(
            """
            CREATE INDEX IF NOT EXISTS idx_transporter_id_entries_dsp_key
            ON transporter_id_entries (dsp_key)
            """
        )
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS associate_data_entries (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                dsp_key TEXT NOT NULL,
                raw_payload TEXT NOT NULL,
                updated_at TEXT NOT NULL
            )
            """
        )
        conn.execute(
            """
            CREATE INDEX IF NOT EXISTS idx_associate_data_entries_dsp_key
            ON associate_data_entries (dsp_key)
            """
        )
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS vehicle_data_entries (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                dsp_key TEXT NOT NULL,
                raw_payload TEXT NOT NULL,
                updated_at TEXT NOT NULL
            )
            """
        )
        conn.execute(
            """
            CREATE INDEX IF NOT EXISTS idx_vehicle_data_entries_dsp_key
            ON vehicle_data_entries (dsp_key)
            """
        )
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS activity_logs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                iso_time TEXT NOT NULL,
                timestamp TEXT,
                action TEXT NOT NULL,
                details TEXT,
                account_key TEXT,
                account_name TEXT,
                session_id TEXT,
                current_tab TEXT,
                page TEXT,
                user_agent TEXT,
                ip_address TEXT,
                raw_payload TEXT,
                created_at TEXT NOT NULL
            )
            """
        )
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS portkey_attendance_report (
                id INTEGER PRIMARY KEY CHECK (id = 1),
                raw_payload TEXT NOT NULL,
                updated_at TEXT NOT NULL
            )
            """
        )
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS user_accounts (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                username TEXT NOT NULL UNIQUE,
                display_name TEXT NOT NULL,
                password_hash TEXT NOT NULL,
                dsp_key TEXT NOT NULL,
                dsp_keys TEXT,
                profile_picture TEXT,
                is_active INTEGER NOT NULL DEFAULT 1,
                must_change_password INTEGER NOT NULL DEFAULT 0,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL,
                last_login_at TEXT
            )
            """
        )
        conn.execute(
            "CREATE INDEX IF NOT EXISTS idx_user_accounts_username ON user_accounts (username)"
        )
        columns = {
            row["name"]
            for row in conn.execute("PRAGMA table_info(user_accounts)").fetchall()
        }
        if "dsp_keys" not in columns:
            conn.execute("ALTER TABLE user_accounts ADD COLUMN dsp_keys TEXT")
            conn.execute("UPDATE user_accounts SET dsp_keys = dsp_key WHERE dsp_keys IS NULL OR dsp_keys = ''")
        if "profile_picture" not in columns:
            conn.execute("ALTER TABLE user_accounts ADD COLUMN profile_picture TEXT")
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS account_audit_logs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER,
                username TEXT NOT NULL,
                action TEXT NOT NULL,
                details TEXT,
                actor TEXT,
                ip_address TEXT,
                created_at TEXT NOT NULL
            )
            """
        )
        conn.execute(
            """
            CREATE INDEX IF NOT EXISTS idx_account_audit_logs_user_id
            ON account_audit_logs (user_id)
            """
        )
        conn.commit()


def extract_initial_phone_lists() -> dict[str, list[dict]]:
    html = HTML_FILE.read_text(encoding="utf-8")
    match = re.search(r"const PHONE_LISTS = \{(.*?)\};\s*\nconst ASSOCIATE_DATA_FALLBACK", html, re.DOTALL)
    if not match:
        return {}

    object_literal = "{" + match.group(1) + "}"
    object_literal = re.sub(r"(\{|,)\s*([A-Za-z_][A-Za-z0-9_]*)\s*:", r'\1 "\2":', object_literal)
    try:
        parsed = json.loads(object_literal)
    except json.JSONDecodeError:
        return {}

    return {key: parsed.get(key, []) for key in PHONE_LIST_DSP_KEYS if isinstance(parsed.get(key), list)}


def seed_phone_lists_if_empty() -> None:
    init_db()
    with get_db_connection() as conn:
        initial_lists = extract_initial_phone_lists()
        now = datetime.now(timezone.utc).isoformat()
        rows_to_insert = []
        for dsp_key, entries in initial_lists.items():
            existing_count = conn.execute(
                "SELECT COUNT(*) AS count FROM phone_list_entries WHERE dsp_key = ?",
                (dsp_key,),
            ).fetchone()["count"]
            if existing_count:
                continue

            for entry in entries:
                rows_to_insert.append(
                    (
                        dsp_key,
                        str(entry.get("label", "")),
                        str(entry.get("lastName", "")),
                        str(entry.get("workPhone", "")),
                        str(entry.get("homePhone", "")),
                        str(entry.get("mobilePhone", "")),
                        now,
                    )
                )

        if rows_to_insert:
            conn.executemany(
                """
                INSERT INTO phone_list_entries (
                    dsp_key, label, last_name, work_phone, home_phone, mobile_phone, updated_at
                )
                VALUES (?, ?, ?, ?, ?, ?, ?)
                """,
                rows_to_insert,
            )
            conn.commit()


def get_client_ip() -> str:
    forwarded_for = request.headers.get("X-Forwarded-For", "")
    if forwarded_for:
        return forwarded_for.split(",")[0].strip()
    return request.headers.get("X-Real-IP", request.remote_addr or "")


def row_to_log(row: sqlite3.Row) -> dict:
    return {
        "id": row["id"],
        "isoTime": row["iso_time"],
        "timestamp": row["timestamp"],
        "action": row["action"],
        "details": row["details"] or "",
        "accountKey": row["account_key"] or "",
        "accountName": row["account_name"] or "",
        "sessionId": row["session_id"] or "",
        "currentTab": row["current_tab"] or "",
        "page": row["page"] or "",
        "userAgent": row["user_agent"] or "",
        "ipAddress": row["ip_address"] or "",
    }


def owner_password_is_valid() -> bool:
    provided = request.headers.get("X-Owner-Password", "")
    return provided == OWNER_PASSWORD


def user_can_access_dsp(dsp_key: str) -> bool:
    if owner_password_is_valid():
        return True
    dsp_keys = session.get("dsp_keys")
    if isinstance(dsp_keys, list):
        return dsp_key in dsp_keys
    return str(session.get("dsp_key") or "") == dsp_key


def validate_dsp_key(dsp_key: str) -> bool:
    return dsp_key in DSP_NAMES


def normalize_dsp_keys(value) -> list[str]:
    if isinstance(value, str):
        try:
            parsed = json.loads(value)
            if isinstance(parsed, list):
                value = parsed
            else:
                value = [value]
        except json.JSONDecodeError:
            value = re.split(r"[,;\s]+", value)
    if not isinstance(value, list):
        value = []

    normalized = []
    for item in value:
        key = str(item or "").strip().lower()
        if key in DSP_NAMES and key not in normalized:
            normalized.append(key)
    return normalized


def serialize_dsp_keys(dsp_keys: list[str]) -> str:
    return json.dumps(dsp_keys, ensure_ascii=True)


def validate_phone_list_dsp(dsp_key: str) -> bool:
    return dsp_key in PHONE_LIST_DSP_KEYS


def validate_storage_dsp(dsp_key: str) -> bool:
    return dsp_key in PHONE_LIST_DSP_KEYS


def row_to_phone_list_entry(row: sqlite3.Row) -> dict:
    return {
        "label": row["label"] or "",
        "lastName": row["last_name"] or "",
        "workPhone": row["work_phone"] or "",
        "homePhone": row["home_phone"] or "",
        "mobilePhone": row["mobile_phone"] or "",
    }


def row_to_transporter_id_entry(row: sqlite3.Row) -> dict:
    return {
        "daKey": row["da_key"] or "",
        "daName": row["da_name"] or "",
        "transporterId": row["transporter_id"] or "",
        "notes": row["notes"] or "",
        "updatedAt": row["updated_at"] or "",
    }


def normalize_username(username: str) -> str:
    return re.sub(r"\s+", "", str(username or "").strip().lower())


def row_to_user_account(row: sqlite3.Row) -> dict:
    dsp_keys = normalize_dsp_keys(row["dsp_keys"] if "dsp_keys" in row.keys() else "")
    if not dsp_keys:
        dsp_keys = normalize_dsp_keys(row["dsp_key"])
    primary_dsp = dsp_keys[0] if dsp_keys else (row["dsp_key"] or "")
    return {
        "id": row["id"],
        "username": row["username"] or "",
        "displayName": row["display_name"] or "",
        "dspKey": primary_dsp,
        "dspKeys": dsp_keys,
        "dspName": DSP_NAMES.get(primary_dsp, primary_dsp or ""),
        "dspNames": [DSP_NAMES.get(key, key) for key in dsp_keys],
        "profilePicture": row["profile_picture"] if "profile_picture" in row.keys() else "",
        "isActive": bool(row["is_active"]),
        "mustChangePassword": bool(row["must_change_password"]),
        "createdAt": row["created_at"] or "",
        "updatedAt": row["updated_at"] or "",
        "lastLoginAt": row["last_login_at"] or "",
    }


def record_account_audit(
    conn: sqlite3.Connection,
    user_id: int | None,
    username: str,
    action: str,
    details: str = "",
    actor: str = "admin",
) -> None:
    conn.execute(
        """
        INSERT INTO account_audit_logs (
            user_id, username, action, details, actor, ip_address, created_at
        )
        VALUES (?, ?, ?, ?, ?, ?, ?)
        """,
        (
            user_id,
            username,
            action,
            details,
            actor,
            get_client_ip()[:80],
            datetime.now(timezone.utc).isoformat(),
        ),
    )


def row_to_account_audit(row: sqlite3.Row) -> dict:
    return {
        "id": row["id"],
        "userId": row["user_id"],
        "username": row["username"] or "",
        "action": row["action"] or "",
        "details": row["details"] or "",
        "actor": row["actor"] or "",
        "ipAddress": row["ip_address"] or "",
        "createdAt": row["created_at"] or "",
    }


def find_associate_data_file(dsp_key: str) -> Path | None:
    for path in ASSOCIATE_DATA_FILES.get(dsp_key, ()):
        if path.exists():
            return path
    return None


def find_vehicle_data_file(dsp_key: str) -> Path | None:
    for path in VEHICLE_DATA_FILES.get(dsp_key, ()):
        if path.exists():
            return path
    return None


def normalize_vehicle_entry(entry: dict) -> dict[str, str]:
    if not isinstance(entry, dict):
        return {}

    normalized = {str(key or "").strip().lower(): value for key, value in entry.items()}

    def value(*keys: str) -> str:
        for key in keys:
            if key in normalized:
                return str(normalized[key] or "").strip()
        return ""

    return {
        "vin": value("vin", "vehicleidentificationnumber"),
        "vehicleName": value("vehiclename", "vehicle name", "vehicle", "name"),
        "stationCode": value("stationcode", "station code", "station"),
        "licensePlateNumber": value(
            "licenseplatenumber",
            "license plate number",
            "license plate",
            "licenseplate",
            "plate",
            "platenumber",
            "plate number"
        ),
        "serviceType": value("servicetype", "service type"),
        "operationalStatus": value("operationalstatus", "operational status"),
        "status": value("status", "operationalstatus", "operational status"),
        "year": value("year"),
        "make": value("make"),
        "model": value("model"),
    }


def format_vehicle_asset_type(entry: dict) -> str | None:
    if not isinstance(entry, dict):
        return None

    year = str(entry.get("year") or "").strip()
    make = str(entry.get("make") or "").strip()
    model = str(entry.get("model") or "").strip()
    if year and make and model:
        return f"{year} {make}, {model}"
    if make and model:
        return f"{make}, {model}"
    if year and make:
        return f"{year} {make}"
    vehicle_name = str(entry.get("vehicleName") or entry.get("vehicle_name") or "").strip()
    if vehicle_name and not vehicle_name.startswith("PRKL-"):
        return vehicle_name
    return None


def find_vehicle_name_by_vin_or_plate(vin: str | None = None, plate: str | None = None) -> str | None:
    vin_value = str(vin or "").strip()
    plate_value = str(plate or "").strip()
    if not vin_value and not plate_value:
        return None

    init_db()
    with get_db_connection() as conn:
        rows = conn.execute(
            "SELECT raw_payload FROM vehicle_data_entries ORDER BY id DESC"
        ).fetchall()

    for row in rows:
        try:
            entries = json.loads(row["raw_payload"])
        except (json.JSONDecodeError, TypeError):
            continue
        for entry in entries:
            if not isinstance(entry, dict):
                continue
            entry = normalize_vehicle_entry(entry)
            entry_vin = entry.get("vin", "").strip()
            entry_plate = entry.get("licensePlateNumber", "").strip()
            if vin_value and entry_vin and vin_value.upper() == entry_vin.upper():
                asset_type = format_vehicle_asset_type(entry)
                if asset_type:
                    return asset_type
            if not vin_value and plate_value and entry_plate and plate_value.upper() == entry_plate.upper():
                asset_type = format_vehicle_asset_type(entry)
                if asset_type:
                    return asset_type

    for dsp_key in VEHICLE_DATA_FILES:
        data_file = find_vehicle_data_file(dsp_key)
        if not data_file:
            continue
        try:
            import openpyxl
        except ImportError:
            continue

        wb = openpyxl.load_workbook(data_file, data_only=True, read_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        headers = [str(header or "").strip() for header in rows[0]]
        for values in rows[1:]:
            row = normalize_vehicle_entry(dict(zip(headers, values)))
            entry_vin = row.get("vin", "").strip()
            entry_plate = row.get("licensePlateNumber", "").strip()
            if vin_value and entry_vin and vin_value.upper() == entry_vin.upper():
                asset_type = format_vehicle_asset_type(row)
                if asset_type:
                    return asset_type
            if not vin_value and plate_value and entry_plate and plate_value.upper() == entry_plate.upper():
                asset_type = format_vehicle_asset_type(row)
                if asset_type:
                    return asset_type

    return None


@app.route("/")
def index():
    return render_template_string(HTML_FILE.read_text(encoding="utf-8"))


@app.route("/api/auth/login", methods=["POST"])
def login_user_account():
    init_db()
    payload = request.get_json(silent=True) or {}
    username = normalize_username(payload.get("username", ""))
    password = str(payload.get("password") or "")
    if not username or not password:
        return jsonify({"error": "username and password are required"}), 400

    with get_db_connection() as conn:
        row = conn.execute(
            """
            SELECT id, username, display_name, password_hash, dsp_key, dsp_keys, profile_picture, is_active,
                   must_change_password, created_at, updated_at, last_login_at
            FROM user_accounts
            WHERE username = ?
            """,
            (username,),
        ).fetchone()

        if not row or not check_password_hash(row["password_hash"], password):
            if row:
                record_account_audit(conn, row["id"], row["username"], "Failed Login", "Incorrect password", "system")
                conn.commit()
            return jsonify({"error": "invalid username or password"}), 401

        if not row["is_active"]:
            record_account_audit(conn, row["id"], row["username"], "Blocked Login", "Inactive account", "system")
            conn.commit()
            return jsonify({"error": "account is inactive"}), 403

        now = datetime.now(timezone.utc).isoformat()
        conn.execute("UPDATE user_accounts SET last_login_at = ? WHERE id = ?", (now, row["id"]))
        record_account_audit(conn, row["id"], row["username"], "Login", "User signed in", "user")
        conn.commit()

        updated = dict(row)
        updated["last_login_at"] = now
        dsp_keys = normalize_dsp_keys(row["dsp_keys"]) or normalize_dsp_keys(row["dsp_key"])
        session["user_id"] = row["id"]
        session["username"] = row["username"]
        session["dsp_key"] = dsp_keys[0] if dsp_keys else row["dsp_key"]
        session["dsp_keys"] = dsp_keys

    return jsonify({"user": row_to_user_account(updated), "source": "server"})


@app.route("/api/auth/change-password", methods=["POST"])
def change_user_password():
    init_db()
    payload = request.get_json(silent=True) or {}
    username = normalize_username(payload.get("username", ""))
    current_password = str(payload.get("currentPassword") or "")
    new_password = str(payload.get("newPassword") or "")
    if not username or not current_password or not new_password:
        return jsonify({"error": "username, current password, and new password are required"}), 400
    if len(new_password) < 8:
        return jsonify({"error": "new password must be at least 8 characters"}), 400

    with get_db_connection() as conn:
        row = conn.execute(
            """
            SELECT id, username, password_hash
            FROM user_accounts
            WHERE username = ? AND is_active = 1
            """,
            (username,),
        ).fetchone()
        if not row or not check_password_hash(row["password_hash"], current_password):
            return jsonify({"error": "current password is incorrect"}), 401

        now = datetime.now(timezone.utc).isoformat()
        conn.execute(
            """
            UPDATE user_accounts
            SET password_hash = ?, must_change_password = 0, updated_at = ?
            WHERE id = ?
            """,
            (generate_password_hash(new_password), now, row["id"]),
        )
        record_account_audit(
            conn,
            row["id"],
            row["username"],
            "Password Changed",
            "User changed password. New password is not stored or visible; admin can reset it if needed.",
            "user",
        )
        conn.commit()

    return jsonify({"ok": True})


@app.route("/api/auth/logout", methods=["POST"])
def logout_user_account():
    session.clear()
    return jsonify({"ok": True})


@app.route("/api/admin/users", methods=["GET"])
def get_user_accounts():
    if not owner_password_is_valid():
        return jsonify({"error": "unauthorized"}), 401
    init_db()
    with get_db_connection() as conn:
        user_rows = conn.execute(
            """
            SELECT id, username, display_name, dsp_key, dsp_keys, profile_picture, is_active, must_change_password,
                   created_at, updated_at, last_login_at
            FROM user_accounts
            ORDER BY username ASC
            """
        ).fetchall()
        audit_rows = conn.execute(
            """
            SELECT id, user_id, username, action, details, actor, ip_address, created_at
            FROM account_audit_logs
            ORDER BY id DESC
            LIMIT 300
            """
        ).fetchall()
    return jsonify({
        "users": [row_to_user_account(row) for row in user_rows],
        "audits": [row_to_account_audit(row) for row in audit_rows],
        "dsps": [{"key": key, "name": name} for key, name in DSP_NAMES.items()],
    })


@app.route("/api/admin/users", methods=["POST"])
def create_user_account():
    if not owner_password_is_valid():
        return jsonify({"error": "unauthorized"}), 401
    init_db()
    payload = request.get_json(silent=True) or {}
    username = normalize_username(payload.get("username", ""))
    display_name = str(payload.get("displayName") or username).strip()
    password = str(payload.get("password") or "")
    dsp_keys = normalize_dsp_keys(payload.get("dspKeys"))
    if not dsp_keys:
        dsp_keys = normalize_dsp_keys(payload.get("dspKey"))
    dsp_key = dsp_keys[0] if dsp_keys else ""
    must_change = bool(payload.get("mustChangePassword", True))
    if not username or not display_name or not password or not dsp_keys:
        return jsonify({"error": "username, display name, password, and at least one DSP are required"}), 400
    if len(password) < 8:
        return jsonify({"error": "password must be at least 8 characters"}), 400

    now = datetime.now(timezone.utc).isoformat()
    try:
        with get_db_connection() as conn:
            cursor = conn.execute(
                """
                INSERT INTO user_accounts (
                    username, display_name, password_hash, dsp_key, dsp_keys, is_active,
                    must_change_password, created_at, updated_at
                )
                VALUES (?, ?, ?, ?, ?, 1, ?, ?, ?)
                """,
                (
                    username,
                    display_name,
                    generate_password_hash(password),
                    dsp_key,
                    serialize_dsp_keys(dsp_keys),
                    int(must_change),
                    now,
                    now,
                ),
            )
            record_account_audit(
                conn,
                cursor.lastrowid,
                username,
                "Account Created",
                "Assigned DSPs: " + ", ".join(DSP_NAMES.get(key, key) for key in dsp_keys),
                "admin",
            )
            conn.commit()
    except sqlite3.IntegrityError:
        return jsonify({"error": "username already exists"}), 409

    return jsonify({"ok": True, "id": cursor.lastrowid}), 201


@app.route("/api/admin/users/<int:user_id>", methods=["PATCH"])
def update_user_account(user_id: int):
    if not owner_password_is_valid():
        return jsonify({"error": "unauthorized"}), 401
    init_db()
    payload = request.get_json(silent=True) or {}
    dsp_keys = normalize_dsp_keys(payload.get("dspKeys"))
    if not dsp_keys:
        dsp_keys = normalize_dsp_keys(payload.get("dspKey"))
    dsp_key = dsp_keys[0] if dsp_keys else ""
    display_name = str(payload.get("displayName") or "").strip()
    is_active = bool(payload.get("isActive", True))
    profile_picture = str(payload.get("profilePicture") or "").strip() if "profilePicture" in payload else None
    if not display_name or not dsp_keys:
        return jsonify({"error": "display name and at least one DSP are required"}), 400
    if profile_picture and not profile_picture.startswith("data:image/"):
        return jsonify({"error": "profile picture must be an image data URL"}), 400

    now = datetime.now(timezone.utc).isoformat()
    with get_db_connection() as conn:
        row = conn.execute("SELECT id, username FROM user_accounts WHERE id = ?", (user_id,)).fetchone()
        if not row:
            return jsonify({"error": "user not found"}), 404
        if profile_picture is None:
            conn.execute(
                """
                UPDATE user_accounts
                SET display_name = ?, dsp_key = ?, dsp_keys = ?, is_active = ?, updated_at = ?
                WHERE id = ?
                """,
                (display_name, dsp_key, serialize_dsp_keys(dsp_keys), int(is_active), now, user_id),
            )
        else:
            conn.execute(
                """
                UPDATE user_accounts
                SET display_name = ?, dsp_key = ?, dsp_keys = ?, profile_picture = ?, is_active = ?, updated_at = ?
                WHERE id = ?
                """,
                (display_name, dsp_key, serialize_dsp_keys(dsp_keys), profile_picture, int(is_active), now, user_id),
            )
        record_account_audit(
            conn,
            user_id,
            row["username"],
            "Account Updated",
            "DSPs: " + ", ".join(DSP_NAMES.get(key, key) for key in dsp_keys) + f" | Active: {is_active}" + (" | Profile picture updated" if profile_picture is not None else ""),
            "admin",
        )
        conn.commit()

    return jsonify({"ok": True})


@app.route("/api/admin/users/<int:user_id>/reset-password", methods=["POST"])
def reset_user_password(user_id: int):
    if not owner_password_is_valid():
        return jsonify({"error": "unauthorized"}), 401
    init_db()
    payload = request.get_json(silent=True) or {}
    new_password = str(payload.get("password") or "")
    if len(new_password) < 8:
        return jsonify({"error": "password must be at least 8 characters"}), 400

    now = datetime.now(timezone.utc).isoformat()
    with get_db_connection() as conn:
        row = conn.execute("SELECT id, username FROM user_accounts WHERE id = ?", (user_id,)).fetchone()
        if not row:
            return jsonify({"error": "user not found"}), 404
        conn.execute(
            """
            UPDATE user_accounts
            SET password_hash = ?, must_change_password = 1, updated_at = ?
            WHERE id = ?
            """,
            (generate_password_hash(new_password), now, user_id),
        )
        record_account_audit(
            conn,
            user_id,
            row["username"],
            "Password Reset",
            "Admin reset password and required the user to change it on next login.",
            "admin",
        )
        conn.commit()

    return jsonify({"ok": True})


@app.route("/api/activity-logs", methods=["GET"])
def get_activity_logs():
    if not owner_password_is_valid():
        return jsonify({"error": "unauthorized"}), 401
    init_db()
    with get_db_connection() as conn:
        rows = conn.execute(
            """
            SELECT id, iso_time, timestamp, action, details, account_key, account_name,
                   session_id, current_tab, page, user_agent, ip_address
            FROM activity_logs
            ORDER BY id ASC
            """
        ).fetchall()
    return jsonify({"logs": [row_to_log(row) for row in rows], "source": "server"})


@app.route("/api/portkey/attendance", methods=["GET"])
def get_portkey_attendance_report():
    if not user_can_access_dsp("portkey"):
        return jsonify({"error": "unauthorized"}), 401
    init_db()
    with get_db_connection() as conn:
        row = conn.execute(
            "SELECT raw_payload, updated_at FROM portkey_attendance_report WHERE id = 1"
        ).fetchone()
    if not row:
        return jsonify({"data": None, "updatedAt": ""})
    try:
        data = json.loads(row["raw_payload"])
    except json.JSONDecodeError:
        data = None
    return jsonify({"data": data, "updatedAt": row["updated_at"] or ""})


@app.route("/api/portkey/attendance", methods=["PUT"])
def save_portkey_attendance_report():
    if not user_can_access_dsp("portkey"):
        return jsonify({"error": "unauthorized"}), 401
    init_db()
    payload = request.get_json(silent=True) or {}
    data = payload.get("data")
    if not isinstance(data, dict):
        return jsonify({"error": "attendance data is required"}), 400
    now = datetime.now(timezone.utc).isoformat()
    raw_payload = json.dumps(data, ensure_ascii=True)
    with get_db_connection() as conn:
        conn.execute(
            """
            INSERT INTO portkey_attendance_report (id, raw_payload, updated_at)
            VALUES (1, ?, ?)
            ON CONFLICT(id) DO UPDATE SET raw_payload = excluded.raw_payload, updated_at = excluded.updated_at
            """,
            (raw_payload, now),
        )
        conn.commit()
    return jsonify({"ok": True, "updatedAt": now})


@app.route("/api/phone-lists/<dsp_key>", methods=["GET"])
def get_phone_list(dsp_key: str):
    if not validate_phone_list_dsp(dsp_key):
        return jsonify({"error": "unknown dsp"}), 404
    if not user_can_access_dsp(dsp_key):
        return jsonify({"error": "forbidden"}), 403

    seed_phone_lists_if_empty()
    with get_db_connection() as conn:
        rows = conn.execute(
            """
            SELECT label, last_name, work_phone, home_phone, mobile_phone
            FROM phone_list_entries
            WHERE dsp_key = ?
            ORDER BY id ASC
            """,
            (dsp_key,),
        ).fetchall()

    return jsonify({"dsp": dsp_key, "entries": [row_to_phone_list_entry(row) for row in rows], "source": "server"})


@app.route("/api/phone-lists/<dsp_key>", methods=["POST"])
def replace_phone_list(dsp_key: str):
    if not validate_phone_list_dsp(dsp_key):
        return jsonify({"error": "unknown dsp"}), 404
    if not user_can_access_dsp(dsp_key):
        return jsonify({"error": "forbidden"}), 403

    payload = request.get_json(silent=True) or {}
    entries = payload.get("entries")
    if not isinstance(entries, list):
        return jsonify({"error": "entries must be a list"}), 400

    normalized_entries = []
    for entry in entries:
        if not isinstance(entry, dict):
            return jsonify({"error": "each entry must be an object"}), 400
        normalized_entries.append(
            (
                dsp_key,
                str(entry.get("label", ""))[:255],
                str(entry.get("lastName", ""))[:255],
                str(entry.get("workPhone", ""))[:80],
                str(entry.get("homePhone", ""))[:80],
                str(entry.get("mobilePhone", ""))[:80],
                datetime.now(timezone.utc).isoformat(),
            )
        )

    init_db()
    with get_db_connection() as conn:
        conn.execute("DELETE FROM phone_list_entries WHERE dsp_key = ?", (dsp_key,))
        if normalized_entries:
            conn.executemany(
                """
                INSERT INTO phone_list_entries (
                    dsp_key, label, last_name, work_phone, home_phone, mobile_phone, updated_at
                )
                VALUES (?, ?, ?, ?, ?, ?, ?)
                """,
                normalized_entries,
            )
        conn.commit()

    return jsonify({"ok": True, "dsp": dsp_key, "count": len(normalized_entries)})


@app.route("/api/transporter-ids/<dsp_key>", methods=["GET"])
def get_transporter_ids(dsp_key: str):
    if not validate_storage_dsp(dsp_key):
        return jsonify({"error": "unknown dsp"}), 404
    if not user_can_access_dsp(dsp_key):
        return jsonify({"error": "forbidden"}), 403

    init_db()
    with get_db_connection() as conn:
        rows = conn.execute(
            """
            SELECT da_key, da_name, transporter_id, notes, updated_at
            FROM transporter_id_entries
            WHERE dsp_key = ?
            ORDER BY da_name COLLATE NOCASE ASC
            """,
            (dsp_key,),
        ).fetchall()

    return jsonify({"dsp": dsp_key, "entries": [row_to_transporter_id_entry(row) for row in rows], "source": "server"})


@app.route("/api/transporter-ids/<dsp_key>", methods=["POST"])
def replace_transporter_ids(dsp_key: str):
    if not validate_storage_dsp(dsp_key):
        return jsonify({"error": "unknown dsp"}), 404
    if not user_can_access_dsp(dsp_key):
        return jsonify({"error": "forbidden"}), 403

    payload = request.get_json(silent=True) or {}
    entries = payload.get("entries")
    if not isinstance(entries, list):
        return jsonify({"error": "entries must be a list"}), 400

    now = datetime.now(timezone.utc).isoformat()
    normalized_entries = []
    for entry in entries:
        if not isinstance(entry, dict):
            return jsonify({"error": "each entry must be an object"}), 400

        da_key = str(entry.get("daKey", "")).strip()[:255]
        da_name = str(entry.get("daName", "")).strip()[:255]
        transporter_id = str(entry.get("transporterId", "")).strip()[:255]
        notes = str(entry.get("notes", "")).strip()[:1000]
        if not da_key or not da_name or not transporter_id:
            continue

        normalized_entries.append((dsp_key, da_key, da_name, transporter_id, notes, now))

    init_db()
    with get_db_connection() as conn:
        conn.execute("DELETE FROM transporter_id_entries WHERE dsp_key = ?", (dsp_key,))
        if normalized_entries:
            conn.executemany(
                """
                INSERT INTO transporter_id_entries (
                    dsp_key, da_key, da_name, transporter_id, notes, updated_at
                )
                VALUES (?, ?, ?, ?, ?, ?)
                """,
                normalized_entries,
            )
        conn.commit()

    return jsonify({"ok": True, "dsp": dsp_key, "count": len(normalized_entries)})


@app.route("/api/associate-data/<dsp_key>", methods=["GET"])
def get_associate_data(dsp_key: str):
    if not validate_storage_dsp(dsp_key):
        return jsonify({"error": "unknown dsp"}), 404
    if not user_can_access_dsp(dsp_key):
        return jsonify({"error": "forbidden"}), 403

    init_db()
    with get_db_connection() as conn:
        row = conn.execute(
            """
            SELECT raw_payload
            FROM associate_data_entries
            WHERE dsp_key = ?
            ORDER BY id DESC
            LIMIT 1
            """,
            (dsp_key,),
        ).fetchone()

    if row:
        try:
            entries = json.loads(row["raw_payload"])
        except json.JSONDecodeError:
            entries = []
        return jsonify({"dsp": dsp_key, "entries": entries, "source": "server"})

    data_file = find_associate_data_file(dsp_key)
    if not data_file:
        return jsonify({"dsp": dsp_key, "entries": [], "source": "missing"})

    with data_file.open(newline="", encoding="utf-8-sig") as handle:
        entries = list(csv.DictReader(handle))

    return jsonify({"dsp": dsp_key, "entries": entries, "source": str(data_file)})


@app.route("/api/associate-data/<dsp_key>", methods=["POST"])
def replace_associate_data(dsp_key: str):
    if not validate_storage_dsp(dsp_key):
        return jsonify({"error": "unknown dsp"}), 404
    if not owner_password_is_valid():
        return jsonify({"error": "owner access required"}), 403

    payload = request.get_json(silent=True) or {}
    entries = payload.get("entries")
    if not isinstance(entries, list):
        return jsonify({"error": "entries must be a list"}), 400

    normalized_entries = []
    for entry in entries:
        if not isinstance(entry, dict):
            return jsonify({"error": "each entry must be an object"}), 400
        normalized_entries.append({str(key): str(value or "") for key, value in entry.items()})

    now = datetime.now(timezone.utc).isoformat()
    init_db()
    with get_db_connection() as conn:
        conn.execute("DELETE FROM associate_data_entries WHERE dsp_key = ?", (dsp_key,))
        conn.execute(
            """
            INSERT INTO associate_data_entries (dsp_key, raw_payload, updated_at)
            VALUES (?, ?, ?)
            """,
            (dsp_key, json.dumps(normalized_entries, ensure_ascii=True), now),
        )
        conn.commit()

    return jsonify({"ok": True, "dsp": dsp_key, "count": len(normalized_entries)})


@app.route("/api/vehicle-data/<dsp_key>", methods=["GET"])
def get_vehicle_data(dsp_key: str):
    if not validate_storage_dsp(dsp_key):
        return jsonify({"error": "unknown dsp"}), 404
    if not user_can_access_dsp(dsp_key):
        return jsonify({"error": "forbidden"}), 403

    init_db()
    with get_db_connection() as conn:
        row = conn.execute(
            """
            SELECT raw_payload
            FROM vehicle_data_entries
            WHERE dsp_key = ?
            ORDER BY id DESC
            LIMIT 1
            """,
            (dsp_key,),
        ).fetchone()

    if row:
        try:
            entries = json.loads(row["raw_payload"])
        except json.JSONDecodeError:
            entries = []
        normalized_entries = [normalize_vehicle_entry(entry) for entry in entries if isinstance(entry, dict)]
        return jsonify({"dsp": dsp_key, "entries": normalized_entries, "source": "server"})

    data_file = find_vehicle_data_file(dsp_key)
    if not data_file:
        return jsonify({"dsp": dsp_key, "entries": [], "source": "missing"})

    try:
        import openpyxl
    except ImportError:
        return jsonify({"dsp": dsp_key, "entries": [], "source": "openpyxl missing"})

    wb = openpyxl.load_workbook(data_file, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return jsonify({"dsp": dsp_key, "entries": [], "source": str(data_file)})

    headers = [str(header or "").strip() for header in rows[0]]
    entries = []
    for values in rows[1:]:
        row = normalize_vehicle_entry(dict(zip(headers, values)))
        if not row.get("vehicleName") or not row.get("vin"):
            continue
        entries.append(
            {
                "vehicleName": str(row.get("vehicleName") or ""),
                "vin": str(row.get("vin") or ""),
                "stationCode": str(row.get("stationCode") or ""),
                "licensePlateNumber": str(row.get("licensePlateNumber") or ""),
                "serviceType": str(row.get("serviceType") or ""),
                "operationalStatus": str(row.get("operationalStatus") or ""),
                "status": str(row.get("status") or ""),
                "year": str(row.get("year") or ""),
                "make": str(row.get("make") or ""),
                "model": str(row.get("model") or ""),
            }
        )

    return jsonify({"dsp": dsp_key, "entries": entries, "source": str(data_file)})


@app.route("/api/vehicle-data/<dsp_key>", methods=["POST"])
def replace_vehicle_data(dsp_key: str):
    if not validate_storage_dsp(dsp_key):
        return jsonify({"error": "unknown dsp"}), 404
    if not user_can_access_dsp(dsp_key):
        return jsonify({"error": "forbidden"}), 403

    payload = request.get_json(silent=True) or {}
    entries = payload.get("entries")
    if not isinstance(entries, list):
        return jsonify({"error": "entries must be a list"}), 400

    normalized_entries = []
    for entry in entries:
        if not isinstance(entry, dict):
            return jsonify({"error": "each entry must be an object"}), 400
        entry = normalize_vehicle_entry(entry)
        vehicle_name = str(entry.get("vehicleName") or "").strip()
        vin = str(entry.get("vin") or "").strip()
        if not vehicle_name or not vin:
            continue
        normalized_entries.append(
            {
                "vehicleName": vehicle_name,
                "vin": vin,
                "stationCode": str(entry.get("stationCode") or "").strip(),
                "licensePlateNumber": str(entry.get("licensePlateNumber") or "").strip(),
                "serviceType": str(entry.get("serviceType") or "").strip(),
                "operationalStatus": str(entry.get("operationalStatus") or "").strip(),
                "status": str(entry.get("status") or "").strip(),
                "year": str(entry.get("year") or "").strip(),
                "make": str(entry.get("make") or "").strip(),
                "model": str(entry.get("model") or "").strip(),
            }
        )

    now = datetime.now(timezone.utc).isoformat()
    init_db()
    with get_db_connection() as conn:
        conn.execute("DELETE FROM vehicle_data_entries WHERE dsp_key = ?", (dsp_key,))
        conn.execute(
            """
            INSERT INTO vehicle_data_entries (dsp_key, raw_payload, updated_at)
            VALUES (?, ?, ?)
            """,
            (dsp_key, json.dumps(normalized_entries, ensure_ascii=True), now),
        )
        conn.commit()

    return jsonify({"ok": True, "dsp": dsp_key, "count": len(normalized_entries)})


def replace_docx_text_nodes(xml: str, replacements: dict[int, str]) -> str:
    pattern = re.compile(r"(<w:t(?:\s[^>]*)?>)(.*?)(</w:t>)", re.DOTALL)
    matches = list(pattern.finditer(xml))

    for index in sorted(replacements.keys(), reverse=True):
        if index < 0 or index >= len(matches):
            continue
        match = matches[index]
        replacement = xml_escape(str(replacements[index] or ""), quote=False)
        xml = xml[: match.start(2)] + replacement + xml[match.end(2) :]

    return xml


def generate_dvic_paper_docx(data: dict[str, str]) -> BytesIO:
    template_file = find_dvic_paper_template_file()
    if not template_file:
        raise FileNotFoundError("DVIC paper inspection template was not found.")

    inspection_date = str(data.get("inspectionDate") or "")
    end_time = str(data.get("endTime") or "")
    inspection_type = str(data.get("inspectionType") or "Pre-Trip")
    asset_type = str(data.get("assetType") or "")

    replacements = {
        12: data.get("firstName", ""),
        18: data.get("lastName", ""),
        29: data.get("vin", ""),
        36: data.get("licensePlate", ""),
        38: "",
        46: asset_type,
        47: "",
        48: "",
        49: "",
        50: "",
        62: data.get("mileage", ""),
        76: data.get("station", ""),
        83: inspection_date,
        84: "",
        85: "",
        86: "",
        93: end_time,
        94: "",
        95: "",
        96: "",
        105: inspection_type,
        106: "",
        107: "",
    }

    output = BytesIO()
    with zipfile.ZipFile(template_file, "r") as source:
        with zipfile.ZipFile(output, "w", compression=zipfile.ZIP_DEFLATED) as target:
            for item in source.infolist():
                content = source.read(item.filename)
                if item.filename == "word/document.xml":
                    xml = content.decode("utf-8")
                    content = replace_docx_text_nodes(xml, replacements).encode("utf-8")
                target.writestr(item, content)

    output.seek(0)
    return output


@app.route("/api/dvic-paper-inspection", methods=["POST"])
def create_dvic_paper_inspection():
    payload = request.get_json(silent=True) or {}
    required = ("firstName", "lastName", "vin", "inspectionDate", "endTime", "inspectionType")
    missing = [field for field in required if not str(payload.get(field) or "").strip()]
    if missing:
        return jsonify({"error": "missing required fields: " + ", ".join(missing)}), 400

    normalized_payload = {key: str(value or "").strip() for key, value in payload.items()}
    da_name = normalized_payload.get("daName") or (
        normalized_payload.get("firstName", "") + " " + normalized_payload.get("lastName", "")
    ).strip()
    inspection_date = normalized_payload.get("inspectionDate") or datetime.now().date().isoformat()
    asset_type = find_vehicle_name_by_vin_or_plate(
        normalized_payload.get("vin"), normalized_payload.get("licensePlate")
    )
    if asset_type:
        normalized_payload["assetType"] = asset_type

    filename_base = re.sub(r'[\\/:*?"<>|]+', "_", f"DVIC Paper Inspection - {da_name or 'DA'} {inspection_date}")

    try:
        docx = generate_dvic_paper_docx(normalized_payload)
    except FileNotFoundError as exc:
        return jsonify({"error": str(exc)}), 404

    return send_file(
        docx,
        as_attachment=True,
        download_name=filename_base + ".docx",
        mimetype="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
    )


@app.route("/api/activity-logs", methods=["POST"])
def create_activity_log():
    init_db()
    payload = request.get_json(silent=True) or {}

    action = str(payload.get("action", "")).strip()[:120]
    if not action:
        return jsonify({"error": "action is required"}), 400

    iso_time = str(payload.get("isoTime") or datetime.now(timezone.utc).isoformat())
    timestamp = str(payload.get("timestamp") or "")
    details = str(payload.get("details") or "")[:1000]
    account_key = str(payload.get("accountKey") or "")[:80]
    account_name = str(payload.get("accountName") or "")[:120]
    session_id = str(payload.get("sessionId") or "")[:120]
    current_tab = str(payload.get("currentTab") or "")[:80]
    page = str(payload.get("page") or "")[:200]
    user_agent = str(payload.get("userAgent") or "")[:255]
    ip_address = get_client_ip()[:80]
    created_at = datetime.now(timezone.utc).isoformat()

    with get_db_connection() as conn:
        cursor = conn.execute(
            """
            INSERT INTO activity_logs (
                iso_time, timestamp, action, details, account_key, account_name,
                session_id, current_tab, page, user_agent, ip_address, raw_payload, created_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                iso_time,
                timestamp,
                action,
                details,
                account_key,
                account_name,
                session_id,
                current_tab,
                page,
                user_agent,
                ip_address,
                json.dumps(payload, ensure_ascii=True),
                created_at,
            ),
        )
        conn.commit()

    return jsonify({"ok": True, "id": cursor.lastrowid}), 201


@app.route("/api/activity-logs", methods=["DELETE"])
def delete_activity_logs():
    if not owner_password_is_valid():
        return jsonify({"error": "unauthorized"}), 401
    init_db()
    with get_db_connection() as conn:
        conn.execute("DELETE FROM activity_logs")
        conn.commit()
    return jsonify({"ok": True})


init_db()
seed_phone_lists_if_empty()


if __name__ == "__main__":
    app.run(debug=True)
